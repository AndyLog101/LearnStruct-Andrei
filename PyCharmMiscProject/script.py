import openpyxl
import os
import random
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from datetime import timedelta, datetime


def file_exists(file_path):
    """Verifică dacă un fișier există."""
    return os.path.isfile(file_path)


# Încarcă fișierul Excel existent
wb = openpyxl.load_workbook('output/generated_timetable.xlsx')
ws = wb.active

# Numără rândurile și coloanele
num_rows = ws.max_row
num_columns = ws.max_column


def random_classroom():
    return str(random.randint(1, 25))


# Asociere materii-culori
subject_colors = {}


def get_subject_color(subject):
    if subject not in subject_colors:
        colors = {
            'Limba germană': 'A8DADC',
            'Matematică': 'A8D08D',
            'Limba română': 'FFE156',
            'Fizică': '6C5B7B',
            'Limba engleză': '457B9D',
            'Chimie': 'F1A7B6',
            'Biologie': 'F1C6D9',
            'Educație fizică': '98C5E5',
            'Istorie': 'D9B7A1',
            'Geografie': 'FF9A8B',
            'Educație antreprenorială': 'E76F51',
            'Educație vizuală': 'F4A261',
            'Psihologie': '2A9D8F',
            'Religie': '6A4C93',
            'Informatică': 'E9C46A',
            'TIC': '2E8B57',
            'Cultură germană': 'A1D6E6'
        }
        subject_colors[subject] = colors.get(subject, 'DDDDDD')  # Gri implicit
    return subject_colors[subject]


# Citește fișierul de intrare
def read_schedule(file_path):
    if not file_exists(file_path):
        raise FileNotFoundError(f"Fișierul '{file_path}' nu a fost găsit.")

    metadata = {}
    teachers_map = {}
    schedule = {}
    section = None
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.rstrip()
            if not line: continue
            if line.endswith(':') and not line.startswith(' '):
                section = line[:-1]
            elif section and ':' in line:
                key, val = line.strip().split(':', 1)
                key, val = key.strip(), val.strip()
                if section == 'metadata':
                    metadata[key] = val
                elif section == 'teachers':
                    teachers_map[key] = val
                elif section == 'schedule':
                    schedule[key] = [s for s in val.split(',')]
    return metadata, teachers_map, schedule


# Formatează celula pentru o materie
def format_cell_for_subject(cell, subject, teacher, room):
    text = f"Sala {room}\n\n{subject}\n\n{teacher}"
    cell.value = text
    cell.font = Font(name="Arial", size=14, italic=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.fill = PatternFill(start_color=get_subject_color(subject), fill_type='solid')


# Creează orarul în Excel
def create_schedule_excel(metadata, teachers_map, schedule, output_path):
    try:
        # Prelucrează metadatele
        start_hour = int(metadata['start_hour'])
        lesson_dur = int(metadata['lesson_duration'])
        break_dur = int(metadata['break_duration'])
        max_per = int(metadata['max_periods_per_day'])
        days = [d.strip() for d in metadata['days'].split(',')]

        # Creează un nou workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Orar'

        # Stiluri
        header_font = Font(name="Arial", bold=True, size=12)
        regular_font = Font(name="Arial", size=11)
        italic_font = Font(name="Arial", italic=True, size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        grey_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='999999'),
            right=Side(style='thin', color='999999'),
            top=Side(style='thin', color='999999'),
            bottom=Side(style='thin', color='999999')
        )

        # Antet coloană
        ora_cell = ws.cell(row=1, column=1, value='Ora')
        ora_cell.font = header_font
        ora_cell.alignment = center_align
        ora_cell.fill = grey_fill
        ora_cell.border = border

        for i, day in enumerate(days, start=2):
            cell = ws.cell(row=1, column=i, value=day)
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = grey_fill
            cell.border = border

        # Înălțimea rândurilor
        for row_num in range(1, num_rows + 1):
            if row_num % 2 == 1:
                ws.row_dimensions[row_num].height = 50
            else:
                ws.row_dimensions[row_num].height = 150

        # Intervalele orare și conținutul
        current = datetime.strptime(f"{start_hour:02d}:00", "%H:%M")
        row = 2
        for p in range(max_per):
            start = current
            end = start + timedelta(minutes=lesson_dur)
            interval = f"{start.strftime('%H:%M')}-{end.strftime('%H:%M')}"
            time_cell = ws.cell(row=row, column=1, value=interval)
            time_cell.font = header_font
            time_cell.alignment = center_align
            time_cell.fill = grey_fill
            time_cell.border = border

            # Completează materiile
            for ci, day in enumerate(days, start=2):
                subs = schedule.get(day, [])
                subj = subs[p] if p < len(subs) else ''
                cell = ws.cell(row=row, column=ci)
                cell.font = regular_font
                cell.alignment = center_align
                cell.border = border

                if subj:
                    teacher = teachers_map.get(subj, random.choice(list(teachers_map.values())))
                    room = random_classroom()
                    format_cell_for_subject(cell, subj, teacher, room)

            current = end
            row += 1

            # Adaugă rândul pentru pauză
            if p < max_per - 1:
                bstart = current
                bend = bstart + timedelta(minutes=break_dur)
                bint = f"{bstart.strftime('%H:%M')}-{bend.strftime('%H:%M')}"
                pause_cell = ws.cell(row=row, column=1, value=bint)
                pause_cell.font = italic_font
                pause_cell.alignment = center_align
                pause_cell.fill = grey_fill
                pause_cell.border = border

                for ci in range(2, len(days) + 2):
                    cell = ws.cell(row=row, column=ci, value=f"Pauză {break_dur} min")
                    cell.font = italic_font
                    cell.alignment = center_align
                    cell.fill = grey_fill
                    cell.border = border

                current = bend
                row += 1

        # Lățimea coloanelor
        ws.column_dimensions['A'].width = 20
        for i in range(2, len(days) + 2):
            col_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[col_letter].width = 25

        # Salvare finală
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        try:
            wb.save(output_path)
            print(f"Orarul a fost salvat cu succes în '{output_path}'")
        except PermissionError:
            base_dir = os.path.dirname(output_path)
            base_name = os.path.basename(output_path)
            name, ext = os.path.splitext(base_name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_path = os.path.join(base_dir, f"{name}_{timestamp}{ext}")
            wb.save(new_path)
            print(f"Fișierul original era blocat. Orarul a fost salvat într-o locație alternativă: '{new_path}'")
    except Exception as e:
        print(f"A apărut o eroare neașteptată: {str(e)}")
        raise


# Rulare
random.seed(42)
metadata, teachers_map, schedule = read_schedule("data/timetable.txt")
create_schedule_excel(metadata, teachers_map, schedule, "output/generated_timetable.xlsx")
